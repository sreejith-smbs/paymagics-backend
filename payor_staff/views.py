from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from openpyxl import Workbook
from .models import PaymentTemplate, TemplatePayee
from payors.models import Payee
from .serializers import PaymentTemplateSerializer, TemplatePayeeSerializer
from django.shortcuts import get_object_or_404
from django.forms.models import model_to_dict
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from django.urls import reverse
from rest_framework.pagination import PageNumberPagination
import json
from openpyxl import load_workbook
from django.db import transaction
from datetime import datetime


# -----------------------------PaymentTemplate Views

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def templates(request):
    template_type = request.query_params.get("type")

    if template_type not in ["payment", "payee"]:
        return Response({"error": "Invalid or missing template type"}, status=status.HTTP_400_BAD_REQUEST)

    if request.method == 'GET':
        templates = PaymentTemplate.objects.filter(template_type=template_type).order_by('-id')
        total_count = templates.count()  
        serializer = PaymentTemplateSerializer(templates, many=True)
 
        return Response({
            "total_count": total_count,
            "results": serializer.data
        })

    elif request.method == 'POST':
        try:
            name = request.data.get("name")
            if not name:
                return Response(
                    {"error": "Template name is required"},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
            if PaymentTemplate.objects.filter(name=name).exists():
                return Response(
                    {"error": f"Template with name '{name}' already exists"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Initialize data safely
            data = request.data.copy()
            
            # Get template_type - adjust this based on how you're receiving it
            template_type = request.GET.get('type') or data.get('template_type')
            if not template_type:
                return Response(
                    {"error": "Template type is required"},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
            data["template_type"] = template_type
        
            # Handle field_order safely
            if 'field_order' in data:
                field_order = data.get("field_order")
                if field_order and isinstance(field_order, str):
                    try:
                        data["field_order"] = json.loads(field_order)
                    except json.JSONDecodeError:
                        return Response(
                            {"error": "field_order must be a valid JSON array"},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                # Ensure field_order is a list if provided
                elif field_order and not isinstance(field_order, list):
                    return Response(
                        {"error": "field_order must be a list"},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            serializer = PaymentTemplateSerializer(data=data)
            if serializer.is_valid():
                serializer.save(created_by=request.user)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            return Response(
                {"error": f"An error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def payment_template_detail(request, pk):
    template = get_object_or_404(PaymentTemplate, pk=pk)

    if request.method == 'GET':
        serializer = PaymentTemplateSerializer(template)
        return Response(serializer.data)

    elif request.method == 'PUT':
        serializer = PaymentTemplateSerializer(template, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        template.delete()
        return Response({"detail": "Template deleted successfully."}, status=status.HTTP_204_NO_CONTENT)


# ----------------------------- Add payee to template
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_payees_to_template(request, template_id):
    try:
        template = PaymentTemplate.objects.get(id=template_id)
    except PaymentTemplate.DoesNotExist:
        return Response({"error": "Template not found"}, status=status.HTTP_404_NOT_FOUND)

    if template.template_type != "payment":
        return Response(
            {"error": "Payees can only be added to 'payment' type templates"},
            status=status.HTTP_400_BAD_REQUEST
        )

    payees_data = request.data.get("payees", [])
    batch_name = request.data.get("batch_name")

    if not batch_name:
        batch_name = f"Batch_{timezone.now().strftime('%Y%m%d%H%M')}"

    if TemplatePayee.objects.filter(batch_name=batch_name).exists():
        return Response(
            {"error": f"Batch name '{batch_name}' already exists"},
            status=status.HTTP_400_BAD_REQUEST
        )

    created_payees = []

    for data in payees_data:
        payee_id = data.get("payee_id")
        if not payee_id:
            continue

        try:
            payee = Payee.objects.get(id=payee_id)
        except Payee.DoesNotExist:
            continue

        payee_dict = model_to_dict(payee)

        # Map dynamic fields from Payee model → template
        dynamic_fields_map = template.dynamic_fields or {}
        dynamic_data = {
            header: payee_dict.get(model_field)
            for header, model_field in dynamic_fields_map.items()
            if model_field in payee_dict
        }

        # ✅ Use provided static_fields, fallback to template defaults
        static_data = data.get("static_fields", template.static_fields or {})

        # ✅ Use provided options_data, fallback to template defaults
        options_data = data.get("options_data", template.options or {})

        template_payee = TemplatePayee.objects.create(
            template=template,
            payee=payee,
            dynamic_data=dynamic_data,
            static_data=static_data,
            options_data=options_data,
            batch_name=batch_name
        )
        created_payees.append(template_payee)




    template_serializer = PaymentTemplateSerializer(template)
    template_data = template_serializer.data

    # Remove ordered_fields from template response if it exists
    if 'ordered_fields' in template_data:
        del template_data['ordered_fields']

    response_payees = []
    for template_payee in created_payees:
        # Combine all data sources
        combined_data = {}
        
        # Add payee basic details
        if hasattr(template_payee, 'payee') and template_payee.payee:
            payee = template_payee.payee
            # Get the actual field names from template's dynamic_fields values
            dynamic_fields_map = template.dynamic_fields or {}
            
            # Map payee fields to template field names dynamically
            for template_field, payee_field in dynamic_fields_map.items():
                if hasattr(payee, payee_field):
                    combined_data[template_field] = getattr(payee, payee_field)
        
        # Add dynamic data (this contains the mapped values from above)
        if template_payee.dynamic_data:
            combined_data.update(template_payee.dynamic_data)
        
        # Add static data
        if template_payee.static_data:
            combined_data.update(template_payee.static_data)
        
        # Add options data
        if template_payee.options_data:
            combined_data.update(template_payee.options_data)
        
        # Apply field ordering - include ALL fields but order specified ones first
        field_order = template.field_order or []
        ordered_payee_details = {}
        
        if field_order:
            # First, add fields that are in field_order (in the specified order)
            for field_name in field_order:
                if field_name in combined_data:
                    ordered_payee_details[field_name] = combined_data[field_name]
            
            # Then, add any remaining fields that weren't in field_order
            for field_name, value in combined_data.items():
                if field_name not in ordered_payee_details:
                    ordered_payee_details[field_name] = value
        else:
            # If no field_order, use the original order
            ordered_payee_details = combined_data
        
        # Build the payee response object
        payee_response = {
            "id": template_payee.id,
            "payee_details": ordered_payee_details,
            "added_at": template_payee.added_at,
            "template": template_payee.template.id,
            "payee": template_payee.payee.id
        }
        
        response_payees.append(payee_response)

    return Response({
        "template": template_data,
        "batch_name": batch_name,
        "payees": response_payees
    }, status=status.HTTP_201_CREATED)



# ----------------------------- download excel 

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_batch_excel(request, batch_name):
    payees = TemplatePayee.objects.filter(batch_name=batch_name)

    if not payees.exists():
        return Response({"error": "No payees found for this batch"}, status=404)

    template = payees.first().template

    wb = Workbook()
    ws = wb.active
    ws.title = batch_name

    # Use field_order if available, otherwise fallback to default order
    if template.field_order:
        headers = template.field_order
    else:
        # Fallback to default order: dynamic -> static -> options
        dynamic_headers = list(template.dynamic_fields.keys()) if template.dynamic_fields else []
        static_headers = list(template.static_fields.keys()) if template.static_fields else []
        options_headers = list(template.options.keys()) if template.options else []
        headers = dynamic_headers + static_headers + options_headers

    # Write headers
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Write data rows
    for row_num, tp in enumerate(payees, start=2):
        for col_num, header in enumerate(headers, start=1):
            value = ""
            
            # Check dynamic data
            if tp.dynamic_data and header in tp.dynamic_data:
                value = tp.dynamic_data[header]
            # Check static data
            elif tp.static_data and header in tp.static_data:
                value = tp.static_data[header]
            # Check options data
            elif tp.options_data and header in tp.options_data:
                value = tp.options_data[header]
            # Check payee model fields (for fields like ben_name, ben_code that are mapped via dynamic_fields)
            elif hasattr(tp, 'payee') and tp.payee:
                # Get the actual field name from template's dynamic_fields mapping
                dynamic_fields_map = template.dynamic_fields or {}
                for template_field, model_field in dynamic_fields_map.items():
                    if template_field == header and hasattr(tp.payee, model_field):
                        value = getattr(tp.payee, model_field)
                        break

            # Convert lists or dicts to a string before writing
            if isinstance(value, (list, dict)):
                value = json.dumps(value)

            ws.cell(row=row_num, column=col_num, value=value)

    # Auto-adjust column widths
    for col_num, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_num)
        max_length = max(
            len(str(ws.cell(row=row, column=col_num).value or "")) for row in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = max_length + 5

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{batch_name}.xlsx"'
    wb.save(response)
    return response


# ----------------------------- excel files
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_batches(request):
    batches = (
        TemplatePayee.objects
        .values("batch_name", "template__id", "template__name")
        .distinct()
    )

    total_count = batches.count()  # Total number of unique batches

    data = []
    for batch in batches:
        data.append({
            "batch_name": batch["batch_name"],
            "template_id": batch["template__id"],
            "template_name": batch["template__name"],
            "download_url": request.build_absolute_uri(
                reverse("download_batch_excel", args=[batch["batch_name"]])
            )
        })
    data.reverse()

    paginator = PageNumberPagination()
    paginator.page_size = 15
    paginated_data = paginator.paginate_queryset(data, request)

    response = paginator.get_paginated_response(paginated_data)
    response.data["total_count"] = total_count  # 👈 Added total count
    return response



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def view_batch_excel(request, batch_name):
    payees = TemplatePayee.objects.filter(batch_name=batch_name)
    if not payees.exists():
        return Response({"error": "Batch not found."}, status=status.HTTP_404_NOT_FOUND)

    # All payees in the same batch share one template
    template = payees.first().template  

    payees_data = TemplatePayeeSerializer(payees, many=True).data
    template_data = PaymentTemplateSerializer(template).data

    response_data = {
        "batch_name": batch_name,
        "template": template_data,
        "total_records": len(payees_data),
        "records": payees_data
    }

    return Response(response_data, status=status.HTTP_200_OK)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_batch_excel(request, batch_name):
    data = request.data
    template_id = data.get("template_id")
    records = data.get("records", [])
    new_batch_name = data.get("new_batch_name", batch_name)

    if not template_id:
        return Response({"error": "Template ID required."}, status=400)
    if not isinstance(records, list) or not records:
        return Response({"error": "records must be a non-empty list."}, status=400)

    try:
        template = PaymentTemplate.objects.get(id=template_id)
    except PaymentTemplate.DoesNotExist:
        return Response({"error": "Template not found."}, status=404)

    payees_qs = TemplatePayee.objects.filter(batch_name=batch_name)
    existing_payee_ids = list(payees_qs.values_list("payee_id", flat=True))

    updated_records = []
    created_records = []
    errors = []

    incoming_payee_ids = []

    for rec in records:
        payee_id = rec.get("payee_id")
        static_fields = rec.get("static_fields", {})
        options_selection = rec.get("options_selection", {})

        if not payee_id:
            errors.append({"error": "Missing payee_id in record."})
            continue

        incoming_payee_ids.append(payee_id)

        try:
            payee = Payee.objects.get(id=payee_id)
        except Payee.DoesNotExist:
            errors.append({"payee_id": payee_id, "error": "Invalid payee_id"})
            continue

        if payee_id in existing_payee_ids:
            # Update existing TemplatePayee
            tp = payees_qs.get(payee_id=payee_id)

            # Update dynamic data from template fields
            dynamic_fields_map = template.dynamic_fields or {}
            for header, model_field in dynamic_fields_map.items():
                tp.dynamic_data[header] = getattr(payee, model_field, "")

            # Update static data
            tp.static_data.update(static_fields or template.static_fields or {})

            # Update options data
            if options_selection:
                for key, value in options_selection.items():
                    if key in (template.options or {}):
                        tp.options_data[key] = value
            else:
                tp.options_data.update(template.options or {})

            # Update template reference
            tp.template = template
            tp.save()

            updated_records.append(TemplatePayeeSerializer(tp).data)

        else:
            # Create new TemplatePayee
            payee_dict = model_to_dict(payee)
            dynamic_fields_map = template.dynamic_fields or {}
            dynamic_data = {
                header: payee_dict.get(model_field)
                for header, model_field in dynamic_fields_map.items()
                if model_field in payee_dict
            }

            static_data = static_fields or template.static_fields or {}
            options_data = options_selection or template.options or {}

            new_tp = TemplatePayee.objects.create(
                template=template,
                payee=payee,
                dynamic_data=dynamic_data,
                static_data=static_data,
                options_data=options_data,
                batch_name=batch_name
            )

            created_records.append(TemplatePayeeSerializer(new_tp).data)

    # Delete payees not present in the new records
    payees_to_delete = payees_qs.exclude(payee_id__in=incoming_payee_ids)
    deleted_count = payees_to_delete.count()
    payees_to_delete.delete()

    # Optional batch rename
    if new_batch_name != batch_name:
        TemplatePayee.objects.filter(batch_name=batch_name).update(batch_name=new_batch_name)

    return Response({
        "message": f"Batch '{batch_name}' updated successfully.",
        "updated_count": len(updated_records),
        "new_count": len(created_records),
        "deleted_count": deleted_count,
        "errors": errors,
        "new_batch_name": new_batch_name,
        "updated_records": updated_records,
        "new_records": created_records
    }, status=200)



# @api_view(['PUT'])
# @permission_classes([IsAuthenticated])
# def update_batch_excel(request, batch_name):
#     data = request.data
#     template_id = data.get("template_id")
#     records = data.get("records", [])
#     new_batch_name = data.get("new_batch_name", batch_name)

#     if not template_id:
#         return Response({"error": "Template ID required."}, status=400)
#     if not isinstance(records, list) or not records:
#         return Response({"error": "records must be a non-empty list."}, status=400)

#     try:
#         template = PaymentTemplate.objects.get(id=template_id)
#     except PaymentTemplate.DoesNotExist:
#         return Response({"error": "Template not found."}, status=404)

#     payees_qs = TemplatePayee.objects.filter(batch_name=batch_name)
#     existing_payee_ids = list(payees_qs.values_list("payee_id", flat=True))

#     updated_records = []
#     created_records = []
#     errors = []

#     incoming_payee_ids = []

#     for rec in records:
#         payee_id = rec.get("payee_id")
#         static_fields = rec.get("static_fields", {})
#         options_selection = rec.get("options_selection", {})

#         if not payee_id:
#             errors.append({"error": "Missing payee_id in record."})
#             continue

#         incoming_payee_ids.append(payee_id)

#         try:
#             payee = Payee.objects.get(id=payee_id)
#         except Payee.DoesNotExist:
#             errors.append({"payee_id": payee_id, "error": "Invalid payee_id"})
#             continue

#         if payee_id in existing_payee_ids:
#             # Update existing TemplatePayee
#             tp = payees_qs.get(payee_id=payee_id)

#             # Update dynamic data from template fields
#             dynamic_fields_map = template.dynamic_fields or {}
#             for header, model_field in dynamic_fields_map.items():
#                 tp.dynamic_data[header] = getattr(payee, model_field, "")

#             # Update static data
#             tp.static_data.update(static_fields or template.static_fields or {})

#             # Update options data
#             if options_selection:
#                 for key, value in options_selection.items():
#                     if key in (template.options or {}):
#                         tp.options_data[key] = value
#             else:
#                 tp.options_data.update(template.options or {})

#             # Update template reference
#             tp.template = template
#             tp.save()

#             updated_records.append(TemplatePayeeSerializer(tp).data)

#         else:
#             # Create new TemplatePayee
#             payee_dict = model_to_dict(payee)
#             dynamic_fields_map = template.dynamic_fields or {}
#             dynamic_data = {
#                 header: payee_dict.get(model_field)
#                 for header, model_field in dynamic_fields_map.items()
#                 if model_field in payee_dict
#             }

#             static_data = static_fields or template.static_fields or {}
#             options_data = options_selection or template.options or {}

#             new_tp = TemplatePayee.objects.create(
#                 template=template,
#                 payee=payee,
#                 dynamic_data=dynamic_data,
#                 static_data=static_data,
#                 options_data=options_data,
#                 batch_name=batch_name
#             )

#             created_records.append(TemplatePayeeSerializer(new_tp).data)

#     # Delete payees not present in the new records
#     payees_to_delete = payees_qs.exclude(payee_id__in=incoming_payee_ids)
#     deleted_count = payees_to_delete.count()
#     payees_to_delete.delete()

#     # Update filename if batch name changed
#     if new_batch_name != batch_name:
#         TemplatePayee.objects.filter(batch_name=batch_name).update(batch_name=new_batch_name)
        
#         # Update filename for all records in the batch
#         batch_records = TemplatePayee.objects.filter(batch_name=new_batch_name)
#         timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#         new_filename = f"{new_batch_name}_{timestamp}.xlsx"
        
#         for record in batch_records:
#             record.filename = new_filename
#             record.save()

#     return Response({
#         "message": f"Batch '{batch_name}' updated successfully.",
#         "updated_count": len(updated_records),
#         "new_count": len(created_records),
#         "deleted_count": deleted_count,
#         "errors": errors,
#         "new_batch_name": new_batch_name,
#         "updated_records": updated_records,
#         "new_records": created_records
#     }, status=200)


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def delete_files(request, batch_name):
    if not batch_name:
        return Response({"error": "Batch name is required."}, status=400)

    deleted_count, _ = TemplatePayee.objects.filter(batch_name=batch_name).delete()

    if deleted_count == 0:
        return Response({"message": f"No TemplatePayees found for batch '{batch_name}'."}, status=404)

    return Response({"message": f" file '{batch_name}' deleted."})



# ----------------------------- fetch options of payment template
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def payment_template_options(request, template_id):
    try:
        template = PaymentTemplate.objects.get(id=template_id, template_type="payment")
    except PaymentTemplate.DoesNotExist:
        return Response({"error": "Payee template not found."}, status=404)

    return Response({
        "id": template.id,
        "template_name": template.name,
        "options": template.options or {},
    })


from payors.serializers import PayeeSerializer
 
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def selected_payees(request):
 
    payee_ids = request.data.get("payees", [])   # list of payee IDs
    list_ids = request.data.get("lists", [])     # list of category IDs
 
    # Start with manually selected payees
    all_payee_ids = set(payee_ids)
 
    # Fetch payees linked to any of the categories in lists
    if list_ids:
        category_payees = Payee.objects.filter(
            categories__id__in=list_ids,
            is_active=True
        ).values_list("id", flat=True)
        all_payee_ids.update(category_payees)
 
    # If no payees found, return empty list
    if not all_payee_ids:
        return Response({"count": 0, "results": []})
 
    # Final queryset
    queryset = Payee.objects.filter(id__in=all_payee_ids, is_active=True).distinct()
 
    # Pagination
    paginator = PageNumberPagination()
    paginator.page_size = 15
    paginated_payees = paginator.paginate_queryset(queryset, request)
 
    serializer = PayeeSerializer(paginated_payees, many=True)
    return paginator.get_paginated_response(serializer.data)

from payor_staff.models import PaymentTemplate
 


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def fetch_payees_for_template(request):

    template_id = request.data.get("template_id")
    if not template_id:
        return Response({"error": "template_id is required"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        template = PaymentTemplate.objects.get(id=template_id)
    except PaymentTemplate.DoesNotExist:
        return Response({"error": "Template not found"}, status=status.HTTP_404_NOT_FOUND)

    payee_ids = request.data.get("payees", [])
    list_ids = request.data.get("lists", [])

    # Combine manually selected payees and payees from lists
    all_payee_ids = set(payee_ids)

    if list_ids:
        category_payees = Payee.objects.filter(
            categories__id__in=list_ids,
            is_active=True
        ).values_list("id", flat=True)
        all_payee_ids.update(category_payees)

    if not all_payee_ids:
        return Response({"count": 0, "results": []})

    queryset = Payee.objects.filter(id__in=all_payee_ids, is_active=True).distinct()

    # Pagination
    paginator = PageNumberPagination()
    paginator.page_size = 15
    paginated_payees = paginator.paginate_queryset(queryset, request)

    results = []
    for payee in paginated_payees:
        # First collect all data without ordering
        all_data = {}

        # Dynamic fields from template
        for field_name, model_field in (template.dynamic_fields or {}).items():
            all_data[field_name] = getattr(payee, model_field, None)

        # Static fields from template
        for field_name, value in (template.static_fields or {}).items():
            all_data[field_name] = value

        # Options fields from template
        for field_name, options in (template.options or {}).items():
            all_data[field_name] = options

        # Apply field ordering if available
        payee_data = {}
        if template.field_order:
            # Add fields in the specified order first
            for field_name in template.field_order:
                if field_name in all_data:
                    payee_data[field_name] = all_data[field_name]
            
            # Then add any remaining fields that weren't in field_order
            for field_name, value in all_data.items():
                if field_name not in payee_data:
                    payee_data[field_name] = value
        else:
            # If no field_order, use the original order
            payee_data = all_data

        results.append(payee_data)

    # Build response
    response_data = {
        "template": {
            "id": template.id,
            "name": template.name,
            "template_type": template.template_type,
            "dynamic_fields": template.dynamic_fields or {},
            "static_fields": template.static_fields or {},
            "options": template.options or {},
            "field_order": template.field_order or [],  # Include field_order in response
            "created_at": template.created_at,
            "created_by": template.created_by.id if template.created_by else None,
        },
        "count": queryset.count(),
        "results": results
    }

    return Response(response_data)





@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_batch_payees(request, batch_name):
    """Get batch payees in exact format with template field ordering"""
    try:
        # Get all template payees for this batch
        template_payees = TemplatePayee.objects.filter(
            batch_name=batch_name
        ).select_related('template', 'payee')
        
        if not template_payees.exists():
            return Response({
                "error": f"No payees found for batch: {batch_name}"
            }, status=404)
        
        # Get the template
        template = template_payees.first().template
        field_order = template.field_order or []
        
        results = []
        for template_payee in template_payees:
            payee = template_payee.payee
            
            # Build combined data from all sources
            combined_data = {}
            
            # 1. Map payee fields through template's dynamic_fields
            if template.dynamic_fields:
                for template_field, model_field in template.dynamic_fields.items():
                    if hasattr(payee, model_field):
                        value = getattr(payee, model_field)
                        if value is not None:
                            combined_data[template_field] = value
            
            # 2. Add static data (overwrites mapped values if same field name)
            if template_payee.static_data:
                combined_data.update(template_payee.static_data)
            
            # 3. Add options data (overwrites if same field name)
            if template_payee.options_data:
                combined_data.update(template_payee.options_data)
            
            # Apply STRICT template field ordering
            ordered_payee_details = {}
            for field_name in field_order:
                if field_name in combined_data:
                    ordered_payee_details[field_name] = combined_data[field_name]
            
            # Add the ordered payee details directly (no nested "payee_details")
            results.append(ordered_payee_details)
        
        # Build template data
        template_data = {
            "id": template.id,
            "name": template.name,
            "template_type": template.template_type,
            "field_order": field_order,
            "dynamic_fields": template.dynamic_fields or {},
            "static_fields": template.static_fields or {},
            "options": template.options or {}
        }
        
        return Response({
            "batch_name": batch_name,
            "template": template_data,
            "payee_count": len(results),
            "payees": results  # Direct array of ordered field objects
        })
        
    except Exception as e:
        return Response({
            "error": f"An error occurred: {str(e)}"
        }, status=500)
    






@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_template(request):
    """
    Upload an Excel file and auto-generate a PaymentTemplate.
    Uses openpyxl.
    """
    file = request.FILES.get("file")
    template_name = request.data.get("template_name")
    template_type = request.data.get("template_type", "payment")

    if not file:
        return Response({"error": "Excel file is required"}, status=400)

    if not template_name:
        return Response({"error": "template_name is required"}, status=400)

    try:
        wb = load_workbook(file)
        sheet = wb.active
    except Exception as e:
        return Response({"error": f"Invalid Excel file: {str(e)}"}, status=400)

    # ------------------------------------
    # READ HEADER ROW
    # ------------------------------------
    headers = []
    for cell in sheet[1]:
        if cell.value:
            headers.append(str(cell.value).strip().replace(" ", "_").lower())

    if not headers:
        return Response({"error": "Excel has no headers"}, status=400)

    # ------------------------------------
    # READ ALL DATA ROWS INTO A LIST
    # ------------------------------------
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        row_dict = {}
        for i, col in enumerate(headers):
            row_dict[col] = row[i]
        rows.append(row_dict)

    if not rows:
        return Response({"error": "Excel has no data rows"}, status=400)

    payee_fields = set([f.name for f in Payee._meta.fields])

    dynamic_fields = {}
    static_fields = {}
    option_fields = {}
    field_order = headers.copy()

    # ------------------------------------
    # CLASSIFY EACH COLUMN
    # ------------------------------------
   

    for col in headers:
        values = [row[col] for row in rows if row[col] is not None]
        unique_values = list(set(values))
        unique_count = len(unique_values)

        # CASE 1
        if col in payee_fields:
            dynamic_fields[col] = col
            continue

        # CASE 2 
        if template_type == "payment":

            # Static field (only one value)
            if unique_count == 1:
                static_fields[col] = str(unique_values[0])
                continue

            # Options field (2–5 values)
            if 2 <= unique_count <= 4:
                option_fields[col] = [str(v) for v in unique_values]
                continue

            # More than 5 values → treat as static 
            static_fields[col] = str(rows[0].get(col))
            continue

        # CASE 3 
        else:
            # For non-payment templates → ALL fields treated as static
            static_fields[col] = str(rows[0].get(col)) if unique_values else None
            continue


    # ------------------------------------
    # CREATE TEMPLATE 
    # ------------------------------------
    try:
        with transaction.atomic():

            template = PaymentTemplate.objects.create(
                name=template_name,
                template_type=template_type,
                dynamic_fields=dynamic_fields or None,
                static_fields=static_fields or None,
                options=option_fields or None,
                field_order=field_order,
                created_by=request.user
            )


    except Exception as e:
        return Response({"error": str(e)}, status=400)

    # ------------------------------------
    # RESPONSE
    # ------------------------------------
    return Response({
        "message": "Template created successfully",
        "template_id": template.id,
        "template_name": template.name,
        "dynamic_fields": dynamic_fields,
        "static_fields": static_fields,
        "option_fields": option_fields,
        "field_order": field_order
    }, status=201)



