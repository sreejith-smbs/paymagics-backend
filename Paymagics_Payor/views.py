from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser
from rest_framework.response import Response
from .serializers import *
from .models import *
from Paymagics_Admin.models import UserProfile, UserRole
import random, string
from rest_framework import status
from rest_framework.pagination import PageNumberPagination
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
from io import BytesIO
from django.core.mail import send_mail
import uuid
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from Paymagics_PayorStaff.models import PaymentTemplate
from django.utils import timezone
from django.db.models import Count, Q
from openpyxl import Workbook
from openpyxl.styles import Font



def generate_referral_code(length=6):
    """Generate a random alphanumeric referral code."""
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=length))

@api_view(["POST", "PUT"])
@permission_classes([IsAuthenticated])
def create_or_update_category(request):
            
    payee_ids = request.data.get("payees")  # expecting a list of IDs
    description = request.data.get("description")  # optional

    category = None
    message = ""

    if request.method == "POST":
        # --- POST: strictly create a new category ---
        category_input = request.data.get("category")  
        if not category_input:
            return Response({"error": "Category name is required for creation."}, status=400)

        # Check if category already exists
        if Category.objects.filter(category__iexact=category_input).exists():
            return Response({"error": "Category with this name already exists."}, status=400)

        # Create new category
        category = Category.objects.create(
            category=category_input,
            description=description or "",
            count=0,
            referral_code=generate_referral_code()
        )
        message = "New category created with referral code."

    elif request.method == "PUT":
        # --- PUT: update existing category ---
             
        category_input = request.data.get("id")  
        if not category_input:
            return Response({"error": "Category ID or name required for update."}, status=400)

        try:
            category_id = int(category_input)
            category = Category.objects.filter(id=category_id).first()
            if not category:
                return Response({"error": "Category with this ID not found."}, status=404)
            message = "Category found by ID."
        except (TypeError, ValueError):
            category_name = category_input
            category = Category.objects.filter(category__iexact=category_name).first()
            if not category:
                return Response({"error": "Category not found by name."}, status=404)
            message = "Category found by name."

        new_category_name = request.data.get("category") 
        # Rename if needed
        if new_category_name:
            if Category.objects.filter(category__iexact=new_category_name).exclude(id=category.id).exists():
                return Response({"error": "A category with the new name already exists."}, status=400)
            old_name = category.category
            category.category = new_category_name
            category.save()
            message += f" Renamed from '{old_name}' to '{new_category_name}'."

        # Update description if provided
        if description:
            category.description = description
            category.save()
            message += " Description updated."

    # --- Assign category to payees if provided ---
    newly_assigned_count = 0
    if payee_ids:
        if not isinstance(payee_ids, list):
            return Response({"error": "Payees must be a list of IDs."}, status=400)

        for payee_id in payee_ids:
            try:
                payee = Payee.objects.get(id=payee_id, is_active=True)
            except Payee.DoesNotExist:
                return Response({"error": f"Payee with ID {payee_id} not found or inactive."}, status=404)

            if not payee.categories.filter(id=category.id).exists():
                payee.categories.add(category)
                newly_assigned_count += 1

        if newly_assigned_count > 0:
            category.count += newly_assigned_count
            category.save()
            message += f" Category assigned to {newly_assigned_count} payee(s) and count updated."
        else:
            message += " Category already assigned to all provided payees."

    return Response({
        "id": category.id,
        "category": category.category,
        "description": category.description,
        "count": category.count,
        "referral_code": category.referral_code,
        "message": message
    }, status=200)


#create payee
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_payee(request):
    serializer = CreatePayeeSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # Get current user and profile
    user = request.user
    try:
        payor = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return Response({'error': 'User profile not found.'}, status=status.HTTP_404_NOT_FOUND)

    # Extract validated fields
    ben_code = serializer.validated_data["ben_code"]
    ben_name = serializer.validated_data["ben_name"]

    # Check for duplicate ben_code under the same payor
    if Payee.objects.filter(ben_code=ben_code, payor=payor, is_active=True).exists():
        return Response(
            {"error": f"Payee with ben_code '{ben_code}' already exists for this payor."},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Handle category validation BEFORE creating payee
    category_input = request.data.get("category")
    category = None

    if category_input:
        if isinstance(category_input, int) or str(category_input).isdigit():
            category = Category.objects.filter(id=int(category_input)).first()
            if not category:
                return Response({'error': 'Category not found.'}, status=status.HTTP_404_NOT_FOUND)
        else:
            category, created = Category.objects.get_or_create(category=category_input, defaults={'count': 0})

    referral_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    payee_type = serializer.validated_data.get("payee_type", "DOMESTIC")

    validated_data = serializer.validated_data.copy()

    if validated_data["payee_type"].upper() == "DOMESTIC":
        validated_data["iban"] = None
        validated_data["swift_code"] = None
        validated_data["sort_code"] = None

    elif validated_data["payee_type"].upper() == "INTERNATIONAL":
        validated_data["ifsc"] = None
        validated_data["acc_no"] = None

    payee = Payee.objects.create(
        ben_code=validated_data["ben_code"],
        ben_name=validated_data["ben_name"],
        add1=validated_data["add1"],
        add2=validated_data["add2"],
        city=validated_data["city"],
        state=validated_data["state"],
        zipcode=validated_data["zipcode"],
        contact=validated_data["contact"],
        email=validated_data["email"],
        payee_type=validated_data["payee_type"],
        acc_no=validated_data.get("acc_no"),
        ifsc=validated_data.get("ifsc"),
        iban=validated_data.get("iban"),
        swift_code=validated_data.get("swift_code"),
        sort_code=validated_data.get("sort_code"),
        bank_name=validated_data.get("bank_name"),
        branch=validated_data.get("branch"),
        bank_account_type=validated_data.get("bank_account_type"),
        referralcode=referral_code,
        payor=payor
    )


    # Add category after payee is created
    if category:
        if not payee.categories.filter(id=category.id).exists():
            payee.categories.add(category)

        # Update category count
        category.count = Payee.objects.filter(categories=category, is_active=True).count()
        category.save()

    return Response(PayeeSerializer(payee).data, status=status.HTTP_201_CREATED)


#edit payee
@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def edit_payee(request, pk):
    if not pk:
        return Response({'error': 'Missing payee ID (pk).'}, status=400)

    # --- Fetch Payee ---
    try:
        payee = Payee.objects.get(pk=pk)
    except Payee.DoesNotExist:
        return Response({'error': 'Payee not found.'}, status=404)

    # --- Validate Requesting User ---
    try:
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return Response({'error': 'User profile not found.'}, status=400)

    serializer = UpdatePayeeSerializer(instance=payee, data=request.data, partial=True)
    if not serializer.is_valid():
        return Response(serializer.errors, status=400)

    validated_data = serializer.validated_data

    # --- Handle Payee Type Switch ---
    new_type = validated_data.get("payee_type", payee.payee_type)

    # Case 1: Same type → Normal field check
    if new_type == payee.payee_type:
        if new_type == "DOMESTIC":
            required = ["acc_no", "ifsc"]
            missing = [f for f in required if not validated_data.get(f) and not getattr(payee, f)]
            if missing:
                return Response(
                    {"error": f"Missing fields for DOMESTIC payee: {', '.join(missing)}"},
                    status=400,
                )
        elif new_type == "INTERNATIONAL":
            required = ["iban", "swift_code"]
            missing = [f for f in required if not validated_data.get(f) and not getattr(payee, f)]
            if missing:
                return Response(
                    {"error": f"Missing fields for INTERNATIONAL payee: {', '.join(missing)}"},
                    status=400,
                )

    # Case 2: Type changed → auto clear old fields, don’t error
    elif new_type != payee.payee_type:
        if new_type == "INTERNATIONAL":
            payee.acc_no = None
            payee.ifsc = None
        elif new_type == "DOMESTIC":
            payee.iban = None
            payee.swift_code = None
            payee.sort_code = None
        # ✅ Apply new type immediately
        payee.payee_type = new_type

    else:
        return Response(
            {"error": "Invalid payee_type. Must be 'DOMESTIC' or 'INTERNATIONAL'."},
            status=400,
        )

    # --- Handle Category Update ---
    category_data = request.data.get("category")
    if category_data is not None:
        if isinstance(category_data, int) or str(category_data).isdigit():
            category = Category.objects.filter(id=int(category_data)).first()
            if not category:
                return Response({"error": "Category not found."}, status=400)
        else:
            category, _ = Category.objects.get_or_create(category=category_data)

        # Update category relationships safely
        for old_cat in payee.categories.all():
            old_cat.count = max((old_cat.count or 1) - 1, 0)
            old_cat.save()

        payee.categories.clear()
        payee.categories.add(category)
        category.count = Payee.objects.filter(categories=category, is_active=True).count()
        category.save()

    # --- Apply Field Updates ---
    for attr, value in validated_data.items():
        if attr not in ["categories", "category", "payee_type"]:
            setattr(payee, attr, value)

    payee.save()

    return Response(PayeeSerializer(payee).data, status=200)


#delete payee
@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def delete_payee(request, pk):
    if not pk:
        return Response({'error': 'Missing payee ID (pk).'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        payee = Payee.objects.get(pk=pk)
    except Payee.DoesNotExist:
        return Response({'error': 'Payee not found.'}, status=status.HTTP_404_NOT_FOUND)

    try:
        payor = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return Response({'error': 'Payor profile not found.'}, status=status.HTTP_400_BAD_REQUEST)

    # Decrement category counts for categories assigned to this payee
    categories = payee.categories.all()
    for category in categories:
        category.count = max((category.count or 1) - 1, 0)  # prevent negative count
        category.save()

    # Optional: clear category assignments from payee
    payee.categories.clear()

    # Soft delete
    payee.is_active = False
    payee.save()

    return Response({'message': 'Payee marked as deleted and category counts updated.'}, status=status.HTTP_200_OK)


#view all payee - paginated
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def payee_list(request):
    queryset = Payee.objects.filter(is_active=True)
    total_count = queryset.count()  # total active payees

    paginator = PageNumberPagination()
    paginator.page_size = 15
    result_page = paginator.paginate_queryset(queryset, request)
    serializer = PayeeSerializer(result_page, many=True)

    response = paginator.get_paginated_response(serializer.data)
    response.data["total_count"] = total_count  #  add to paginated response
    return response



#single payee view
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def payee_detail(request, pk):
    try:
        payee = Payee.objects.get(pk=pk, is_active=True)
    except Payee.DoesNotExist:
        return Response({'error': 'Payee not found.'}, status=status.HTTP_404_NOT_FOUND)
    
    serializer = PayeeSerializer(payee)
    return Response(serializer.data, status=status.HTTP_200_OK)


#view all lists 
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def view_list(request):
    queryset = Category.objects.all()
    total_count = queryset.count()  

    serializer = CategorySerializer(queryset, many=True)

    return Response({
        "total_count": total_count,
        "results": serializer.data
    })



#delete category
@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def delete_categ(request, pk):
    if not pk:
        return Response({'error': 'Missing category ID (pk).'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        category = Category.objects.get(pk=pk)
    except Category.DoesNotExist:
        return Response({'error': 'Category not found.'}, status=status.HTTP_404_NOT_FOUND)

    category.delete()

    return Response({'message': 'Category deleted successfully.'}, status=status.HTTP_200_OK)


#payee list export to excel based on template
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def export_payees_excel(request, template_id):
    # Handle GET vs POST inputs
    if request.method == "GET":
        query = request.GET.get("q", "")
        download = request.GET.get("download", "false").lower() == "true"
    else:  # POST
        query = request.data.get("q", "")
        download = str(request.data.get("download", "false")).lower() == "true"

    # Get template (needed only for Excel)
    try:
        template = PaymentTemplate.objects.get(id=template_id)
    except PaymentTemplate.DoesNotExist:
        return HttpResponse("Template not found.", status=404)

    # Base queryset
    payees = Payee.objects.filter(is_active=True)

    # Filtering
    if query:
        matching_categories = Category.objects.filter(category__icontains=query)
        payees = payees.filter(
            Q(ben_code__icontains=query) |
            Q(ben_name__icontains=query) |
            Q(contact__icontains=query) |
            Q(email__icontains=query) |
            Q(categories__in=matching_categories)
        ).distinct()

    if not payees.exists():
        return HttpResponse("No payees found.", status=404)

    # 👉 If not downloading, return paginated JSON (view mode)
    if not download:
        paginator = PageNumberPagination()
        paginator.page_size = 15
        result_page = paginator.paginate_queryset(payees, request)
        serializer = PayeeSerializer(result_page, many=True)
        return paginator.get_paginated_response(serializer.data)

    # 👉 Otherwise, export as Excel
    dynamic_fields = template.dynamic_fields or {}
    static_fields = template.static_fields or {}

    dynamic_headers = list(dynamic_fields.keys())
    static_headers = list(static_fields.keys())
    headers = dynamic_headers + static_headers

    wb = Workbook()
    ws = wb.active
    ws.title = template.name or "Payee Export"

    # Add template name
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=f"Template: {template.name}")
    title_cell.font = Font(bold=True, size=14)

    # Headers
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Data
    for row_num, payee in enumerate(payees, start=3):
        for col_num, header in enumerate(headers, start=1):
            if header in dynamic_fields:
                model_field = dynamic_fields[header]
                value = getattr(payee, model_field, "")
            else:
                value = static_fields.get(header, "")
            ws.cell(row=row_num, column=col_num, value=value)

    # Auto column width
    for col_num, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(col_num)
        max_length = max(
            len(str(ws.cell(row=row, column=col_num).value or "")) 
            for row in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = max_length + 5

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{template.name}_payees_export_{timestamp}.xlsx"

    response = HttpResponse(
        file_stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response




from django.db import transaction


def generate_unique_ben_code():
    """Generate a unique 8-character beneficiary code."""
    while True:
        ben_code = "BEN" + ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
        if not Payee.objects.filter(ben_code=ben_code).exists():
            return ben_code




#view payees corresponding to list
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def payees_in_list(request, category):
    payees = Payee.objects.filter(categories__id=category, is_active=True)
    total_count = payees.count()  # 👈 total number of payees in this category

    if total_count == 0:
        return Response({'error': 'No payees found for this category.'}, status=status.HTTP_404_NOT_FOUND)

    paginator = PageNumberPagination()
    paginator.page_size = 15
    result_page = paginator.paginate_queryset(payees, request)
    serializer = PayeeSerializer(result_page, many=True)

    response = paginator.get_paginated_response(serializer.data)
    response.data["total_count"] = total_count  # 👈 add total count to response

    return response




#dashboard -payee
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def list_counts(request):
    # Total categories
    total_categories = Category.objects.count()

    # Active categories: categories with at least one active payee
    active_categories = Category.objects.annotate(
        active_payees_count=Count('payees', filter=Q(payees__is_active=True))
    ).filter(active_payees_count__gt=0).count()

    # Total payees
    total_payees = Payee.objects.filter(is_active=True).count()

    data = {
        "total_lists": total_categories,
        "active_lists": active_categories,
        "total_payees": total_payees
    }

    return Response(data, status=200)


#------------------------delete file
@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def delete_files(request, batch_name):
    if not batch_name:
        return Response({"error": "Batch name is required."}, status=400)

    deleted_count, _ = TemplatePayee.objects.filter(batch_name=batch_name).delete()

    if deleted_count == 0:
        return Response({"message": f"No TemplatePayees found for batch '{batch_name}'."}, status=404)

    return Response({"message": f" file '{batch_name}' deleted."})



# ----------------------------- fetch options of payment template
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def payment_template_options(request, template_id):
    try:
        template = PaymentTemplate.objects.get(id=template_id, template_type="payment")
    except PaymentTemplate.DoesNotExist:
        return Response({"error": "Payee template not found."}, status=404)

    return Response({
        "id": template.id,
        "template_name": template.name,
        "options": template.options or {},
    })



@api_view(["POST"])
@permission_classes([IsAuthenticated])
def remove_payee_from_category(request):
    payee_id = request.data.get("payee_id")
    category_id = request.data.get("category_id")

    # Validate required fields
    if not payee_id or not category_id:
        return Response(
            {"error": "Both 'payee_id' and 'category_id' are required."},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Fetch payee
    try:
        payee = Payee.objects.get(id=payee_id, is_active=True)
    except Payee.DoesNotExist:
        return Response({"error": "Payee not found or inactive."}, status=status.HTTP_404_NOT_FOUND)

    # Fetch category
    try:
        category = Category.objects.get(id=category_id)
    except Category.DoesNotExist:
        return Response({"error": "Category not found."}, status=status.HTTP_404_NOT_FOUND)

    # Remove relationship if exists
    if payee.categories.filter(id=category.id).exists():
        payee.categories.remove(category)

        # Decrease count safely
        if category.count > 0:
            category.count -= 1
            category.save()

        message = f"Category '{category.category}' removed from payee '{payee.ben_name}'."
    else:
        message = f"Payee '{payee.ben_name}' is not linked to category '{category.category}'."

    return Response({
        "payee_id": payee.id,
        "payee_name": payee.ben_name,
        "category_id": category.id,
        "category_name": category.category,
        "category_count": category.count,
        "message": message
    }, status=status.HTTP_200_OK)







from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.conf import settings
import json



def generate_category_referral_code(category: Category, referrer: UserProfile = None):
    """
    Generates a one-time referral code linked to a specific Category.
    Returns both the code and the referral link fragment.
    """
    referral = CategoryReferralCode.objects.create(category=category, referrer=referrer)
    return {
        "code": referral.code,
        "category": category.category,
        "referrer": referrer.id if referrer else None,
        "created_at": referral.created_at,
    }



def create_category_referral_code(referrer_profile, category_id):
    """
    Helper — Generate a one-time referral link for a Category.
    Returns the referral code string.
    """
    try:
        category = Category.objects.get(id=category_id)
    except Category.DoesNotExist:
        return None

    data = generate_category_referral_code(category, referrer_profile)
    return data["code"]



@csrf_exempt  # only for testing with Postman
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_invite_email(request):
    try:
        # Handle JSON or form POST
        if request.content_type == "application/json":
            data = json.loads(request.body)
        else:
            data = request.POST

        recipient_email = data.get("email")
        custom_msg = data.get("message", "You are invited!")
        category_id = data.get("category")

        if not recipient_email:
            return Response({"status": "error", "message": "Email not provided"}, status=400)
        if not category_id:
            return Response({"status": "error", "message": "Category not provided"}, status=400)

        # ✅ Get logged-in user's profile
        referrer_profile = request.user.profile

        # ✅ Generate referral code (correctly)
        referral_code = create_category_referral_code(referrer_profile, category_id)
        if not referral_code:
            return Response({'error': 'Category not found.'}, status=404)

        # ✅ Build referral link
        frontend_base_url = "https://paymagics-frontend.vercel.app/invite"
        referral_link = f"{frontend_base_url}?referral_code={referral_code}"

        # ✅ Render HTML email
        message = render_to_string('referral_email.html', {
            'referral_link': referral_link,
            'custom_msg': custom_msg
        })

        # ✅ Send email
        email = EmailMessage(
            subject="Invitation to Join",
            body=message,
            from_email=settings.EMAIL_HOST_USER,
            to=[recipient_email]
        )
        email.content_subtype = 'html'
        email.send()

        return Response({"status": "success", "message": f"Invite sent to {recipient_email}"})

    except Exception as e:
        return Response({"status": "error", "message": str(e)}, status=500)



def generate_unique_ben_code(length=8):
    """Generate a unique ben_code."""
    while True:
        ben_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=length))
        if not Payee.objects.filter(ben_code=ben_code).exists():
            return ben_code




@api_view(["POST"])
@permission_classes([AllowAny])
@transaction.atomic
def create_payee_via_referral(request, referral_code):
    """
    Public API — Create a Payee using a one-time Category referral code.
    - No authentication required.
    - Each CategoryReferralCode can only be used once.
    """

    # 1️⃣ Validate and fetch referral record
    try:
        cat_ref = CategoryReferralCode.objects.select_related('category').get(code=referral_code)
    except CategoryReferralCode.DoesNotExist:
        return Response({"error": "Invalid or expired referral code."}, status=status.HTTP_400_BAD_REQUEST)

    if cat_ref.is_used:
        return Response({"error": "This referral link has already been used."}, status=status.HTTP_400_BAD_REQUEST)

    category = cat_ref.category

    # 2️⃣ Validate Payee input
    serializer = CreatePayeeSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    validated_data = serializer.validated_data
    payee_type = validated_data.get("payee_type", "DOMESTIC").upper()

    # 3️⃣ Generate codes
    ben_code = generate_unique_ben_code()
    new_referral_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

    # 4️⃣ Validate banking fields
    acc_no, ifsc, iban, swift = (
        validated_data.get("acc_no"),
        validated_data.get("ifsc"),
        validated_data.get("iban"),
        validated_data.get("swift_code"),
    )

    if payee_type == "DOMESTIC":
        if not acc_no or not ifsc:
            return Response({"error": "Domestic payees need acc_no and ifsc."}, status=400)
        validated_data.update({"iban": None, "swift_code": None, "sort_code": None})
    elif payee_type == "INTERNATIONAL":
        if not iban or not swift:
            return Response({"error": "International payees need iban and swift_code."}, status=400)
        validated_data.update({"acc_no": None, "ifsc": None})
    else:
        return Response({"error": "Invalid payee_type."}, status=400)

    # 5️⃣ Prevent duplicate Payee
    if Payee.objects.filter(email=validated_data["email"], is_active=True).exists():
        return Response({"error": "A Payee with this email already exists."}, status=400)

    # 6️⃣ Create Payee
    payee = Payee.objects.create(
        ben_code=ben_code,
        ben_name=validated_data["ben_name"],
        add1=validated_data.get("add1"),
        add2=validated_data.get("add2"),
        city=validated_data.get("city"),
        state=validated_data.get("state"),
        zipcode=validated_data.get("zipcode"),
        contact=validated_data.get("contact"),
        email=validated_data.get("email"),
        payee_type=payee_type,
        acc_no=validated_data.get("acc_no"),
        ifsc=validated_data.get("ifsc"),
        iban=validated_data.get("iban"),
        swift_code=validated_data.get("swift_code"),
        sort_code=validated_data.get("sort_code"),
        bank_name=validated_data.get("bank_name"),
        branch=validated_data.get("branch"),
        bank_account_type=validated_data.get("bank_account_type"),
        referralcode=new_referral_code,
        payor=cat_ref.referrer  # optional link to inviter
    )

    payee.categories.add(category)

    # 7️⃣ Mark referral as used
    cat_ref.mark_used(payee)

    # 8️⃣ Update category count
    category.count = Payee.objects.filter(categories=category, is_active=True).count()
    category.save()

    # 9️⃣ Return created Payee
    return Response(PayeeSerializer(payee).data, status=status.HTTP_201_CREATED)



