from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from openpyxl import Workbook
from .models import PaymentTemplate, TemplatePayee, PaymentTemp
from Paymagics_Payor.models import Payee
from .serializers import PaymentTemplateSerializer, TemplatePayeeSerializer, PaymentTempSerializer
from django.shortcuts import get_object_or_404
from django.forms.models import model_to_dict
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from django.urls import reverse
from rest_framework.pagination import PageNumberPagination
import json



# -----------------------------PaymentTemplate Views

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def templates(request):
    template_type = request.query_params.get("type")

    if template_type not in ["payment", "payee"]:
        return Response({"error": "Invalid or missing template type"}, status=status.HTTP_400_BAD_REQUEST)

    if request.method == 'GET':
        templates = PaymentTemplate.objects.filter(template_type=template_type)
        total_count = templates.count()  
        serializer = PaymentTemplateSerializer(templates, many=True)
 
        return Response({
            "total_count": total_count,
            "results": serializer.data
        })

    elif request.method == 'POST':
        name = request.data.get("name")
        if PaymentTemplate.objects.filter(name=name).exists():
            return Response(
                {"error": f"Template with name '{name}' already exists"},
                status=status.HTTP_400_BAD_REQUEST
            )

        data = request.data.copy()
        data["template_type"] = template_type   

        serializer = PaymentTemplateSerializer(data=data)
        if serializer.is_valid():
            serializer.save(created_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def payment_template_detail(request, pk):
    template = get_object_or_404(PaymentTemplate, pk=pk)

    if request.method == 'GET':
        serializer = PaymentTemplateSerializer(template)
        return Response(serializer.data)

    elif request.method == 'PUT':
        serializer = PaymentTemplateSerializer(template, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(created_by=request.user)
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        template.delete()
        return Response({"detail": "Template deleted successfully."}, status=status.HTTP_204_NO_CONTENT)


# ----------------------------- Add payee to template
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def add_payees_to_template(request, template_id):
    try:
        template = PaymentTemplate.objects.get(id=template_id)
    except PaymentTemplate.DoesNotExist:
        return Response({"error": "Template not found"}, status=status.HTTP_404_NOT_FOUND)

    if template.template_type != "payment":
        return Response(
            {"error": "Payees can only be added to 'payment' type templates"},
            status=status.HTTP_400_BAD_REQUEST
        )

    payees_data = request.data.get("payees", [])
    batch_name = request.data.get("batch_name")

    if not batch_name:
        batch_name = f"Batch_{timezone.now().strftime('%Y%m%d%H%M')}"

    if TemplatePayee.objects.filter(batch_name=batch_name).exists():
        return Response(
            {"error": f"Batch name '{batch_name}' already exists"},
            status=status.HTTP_400_BAD_REQUEST
        )

    created_payees = []

    for data in payees_data:
        payee_id = data.get("payee_id")
        if not payee_id:
            continue

        try:
            payee = Payee.objects.get(id=payee_id)
        except Payee.DoesNotExist:
            continue

        payee_dict = model_to_dict(payee)

        # Map dynamic fields from Payee model → template
        dynamic_fields_map = template.dynamic_fields or {}
        dynamic_data = {
            header: payee_dict.get(model_field)
            for header, model_field in dynamic_fields_map.items()
            if model_field in payee_dict
        }

        # ✅ Use provided static_fields, fallback to template defaults
        static_data = data.get("static_fields", template.static_fields or {})

        # ✅ Use provided options_data, fallback to template defaults
        options_data = data.get("options_data", template.options or {})

        template_payee = TemplatePayee.objects.create(
            template=template,
            payee=payee,
            dynamic_data=dynamic_data,
            static_data=static_data,
            options_data=options_data,
            batch_name=batch_name
        )
        created_payees.append(template_payee)

    serializer = TemplatePayeeSerializer(created_payees, many=True)
    return Response({
        "template": PaymentTemplateSerializer(template).data,
        "batch_name": batch_name,
        "payees": serializer.data
    }, status=status.HTTP_201_CREATED)



# ----------------------------- download excel 
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_batch_excel(request, batch_name):
    payees = TemplatePayee.objects.filter(batch_name=batch_name)

    if not payees.exists():
        return Response({"error": "No payees found for this batch"}, status=404)

    template = payees.first().template

    wb = Workbook()
    ws = wb.active
    ws.title = batch_name

    dynamic_headers = list(template.dynamic_fields.keys()) if template.dynamic_fields else []
    static_headers = list(template.static_fields.keys()) if template.static_fields else []
    options_headers = list(template.options.keys()) if template.options else []

    headers = dynamic_headers + static_headers + options_headers

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for row_num, tp in enumerate(payees, start=2):
        for col_num, h in enumerate(headers, start=1):
            value = ""
            if tp.dynamic_data and h in tp.dynamic_data:
                value = tp.dynamic_data[h]
            elif tp.static_data and h in tp.static_data:
                value = tp.static_data[h]
            elif tp.options_data and h in tp.options_data:
                value = tp.options_data[h]

            # Convert lists or dicts to a string before writing
            if isinstance(value, (list, dict)):
                value = json.dumps(value)

            ws.cell(row=row_num, column=col_num, value=value)


    for col_num, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_num)
        max_length = max(
            len(str(ws.cell(row=row, column=col_num).value or "")) for row in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = max_length + 5

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{batch_name}.xlsx"'
    wb.save(response)
    return response



# ----------------------------- excel files
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_batches(request):
    batches = (
        TemplatePayee.objects
        .values("batch_name", "template__id", "template__name")
        .distinct()
    )

    total_count = batches.count()  # Total number of unique batches

    data = []
    for batch in batches:
        data.append({
            "batch_name": batch["batch_name"],
            "template_id": batch["template__id"],
            "template_name": batch["template__name"],
            "download_url": request.build_absolute_uri(
                reverse("download_batch_excel", args=[batch["batch_name"]])
            )
        })

    paginator = PageNumberPagination()
    paginator.page_size = 15
    paginated_data = paginator.paginate_queryset(data, request)

    response = paginator.get_paginated_response(paginated_data)
    response.data["total_count"] = total_count  # 👈 Added total count
    return response



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def view_batch_excel(request, batch_name):
    payees = TemplatePayee.objects.filter(batch_name=batch_name)
    if not payees.exists():
        return Response({"error": "Batch not found."}, status=status.HTTP_404_NOT_FOUND)

    # All payees in the same batch share one template
    template = payees.first().template  

    payees_data = TemplatePayeeSerializer(payees, many=True).data
    template_data = PaymentTemplateSerializer(template).data

    response_data = {
        "batch_name": batch_name,
        "template": template_data,
        "total_records": len(payees_data),
        "records": payees_data
    }

    return Response(response_data, status=status.HTTP_200_OK)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_batch_excel(request, batch_name):
    data = request.data
    template_id = data.get("template_id")
    records = data.get("records", [])
    new_batch_name = data.get("new_batch_name", batch_name)

    if not template_id:
        return Response({"error": "Template ID required."}, status=400)
    if not isinstance(records, list) or not records:
        return Response({"error": "records must be a non-empty list."}, status=400)

    try:
        template = PaymentTemplate.objects.get(id=template_id)
    except PaymentTemplate.DoesNotExist:
        return Response({"error": "Template not found."}, status=404)

    payees_qs = TemplatePayee.objects.filter(batch_name=batch_name)
    existing_payee_ids = list(payees_qs.values_list("payee_id", flat=True))

    updated_records = []
    created_records = []
    errors = []

    for rec in records:
        payee_id = rec.get("payee_id")
        static_fields = rec.get("static_fields", {})
        options_selection = rec.get("options_selection", {})

        if not payee_id:
            errors.append({"error": "Missing payee_id in record."})
            continue

        try:
            payee = Payee.objects.get(id=payee_id)
        except Payee.DoesNotExist:
            errors.append({"payee_id": payee_id, "error": "Invalid payee_id"})
            continue

        # If payee already exists in batch → update it
        if payee_id in existing_payee_ids:
            tp = payees_qs.get(payee_id=payee_id)

            if static_fields:
                tp.static_data.update(static_fields)

            if options_selection:
                for key, value in options_selection.items():
                    if key in (template.options or {}):
                        tp.options_data[key] = value

            tp.save()
            updated_records.append(TemplatePayeeSerializer(tp).data)

        # Else → create a new TemplatePayee record
        else:
            payee_dict = model_to_dict(payee)
            dynamic_fields_map = template.dynamic_fields or {}
            dynamic_data = {
                header: payee_dict.get(model_field)
                for header, model_field in dynamic_fields_map.items()
                if model_field in payee_dict
            }

            static_data = static_fields or template.static_fields or {}
            options_data = options_selection or template.options or {}

            new_tp = TemplatePayee.objects.create(
                template=template,
                payee=payee,
                dynamic_data=dynamic_data,
                static_data=static_data,
                options_data=options_data,
                batch_name=batch_name
            )

            created_records.append(TemplatePayeeSerializer(new_tp).data)

    # Optional batch rename
    if new_batch_name != batch_name:
        TemplatePayee.objects.filter(batch_name=batch_name).update(batch_name=new_batch_name)

    return Response({
        "message": f"Batch '{batch_name}' updated successfully.",
        "updated_count": len(updated_records),
        "new_count": len(created_records),
        "errors": errors,
        "new_batch_name": new_batch_name,
        "updated_records": updated_records,
        "new_records": created_records
    }, status=200)


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def delete_files(request, batch_name):
    if not batch_name:
        return Response({"error": "Batch name is required."}, status=400)

    deleted_count, _ = TemplatePayee.objects.filter(batch_name=batch_name).delete()

    if deleted_count == 0:
        return Response({"message": f"No TemplatePayees found for batch '{batch_name}'."}, status=404)

    return Response({"message": f" file '{batch_name}' deleted."})



# ----------------------------- fetch options of payment template
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def payment_template_options(request, template_id):
    try:
        template = PaymentTemplate.objects.get(id=template_id, template_type="payment")
    except PaymentTemplate.DoesNotExist:
        return Response({"error": "Payee template not found."}, status=404)

    return Response({
        "id": template.id,
        "template_name": template.name,
        "options": template.options or {},
    })


from Paymagics_Payor.serializers import PayeeSerializer
 
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def selected_payees(request):
 
    payee_ids = request.data.get("payees", [])   # list of payee IDs
    list_ids = request.data.get("lists", [])     # list of category IDs
 
    # Start with manually selected payees
    all_payee_ids = set(payee_ids)
 
    # Fetch payees linked to any of the categories in lists
    if list_ids:
        category_payees = Payee.objects.filter(
            categories__id__in=list_ids,
            is_active=True
        ).values_list("id", flat=True)
        all_payee_ids.update(category_payees)
 
    # If no payees found, return empty list
    if not all_payee_ids:
        return Response({"count": 0, "results": []})
 
    # Final queryset
    queryset = Payee.objects.filter(id__in=all_payee_ids, is_active=True).distinct()
 
    # Pagination
    paginator = PageNumberPagination()
    paginator.page_size = 15
    paginated_payees = paginator.paginate_queryset(queryset, request)
 
    serializer = PayeeSerializer(paginated_payees, many=True)
    return paginator.get_paginated_response(serializer.data)

from Paymagics_Payor.models import PaymentTemplate
 
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def fetch_payees_for_template(request):
 
    template_id = request.data.get("template_id")
    if not template_id:
        return Response({"error": "template_id is required"}, status=status.HTTP_400_BAD_REQUEST)
 
    try:
        template = PaymentTemplate.objects.get(id=template_id)
    except PaymentTemplate.DoesNotExist:
        return Response({"error": "Template not found"}, status=status.HTTP_404_NOT_FOUND)
 
    payee_ids = request.data.get("payees", [])
    list_ids = request.data.get("lists", [])
 
    # Combine manually selected payees and payees from lists
    all_payee_ids = set(payee_ids)
 
    if list_ids:
        category_payees = Payee.objects.filter(
            categories__id__in=list_ids,
            is_active=True
        ).values_list("id", flat=True)
        all_payee_ids.update(category_payees)
 
    if not all_payee_ids:
        return Response({"count": 0, "results": []})
 
    queryset = Payee.objects.filter(id__in=all_payee_ids, is_active=True).distinct()
 
    # Pagination
    paginator = PageNumberPagination()
    paginator.page_size = 15
    paginated_payees = paginator.paginate_queryset(queryset, request)
 
    results = []
    for payee in paginated_payees:
        payee_data = {}
 
        # Dynamic fields from template
        for field_name, model_field in (template.dynamic_fields or {}).items():
            payee_data[field_name] = getattr(payee, model_field, None)
 
        # Static fields from template
        for field_name, value in (template.static_fields or {}).items():
            payee_data[field_name] = value
 
        # Options fields from template
        for field_name, options in (template.options or {}).items():
            payee_data[field_name] = options
 
        results.append(payee_data)
 
    # Build response
    response_data = {
        "template": {
            "id": template.id,
            "name": template.name,
            "template_type": template.template_type,
            "dynamic_fields": template.dynamic_fields or {},
            "static_fields": template.static_fields or {},
            "options": template.options or {},
            "created_at": template.created_at,
            "created_by": template.created_by.id if template.created_by else None,
        },
        "count": queryset.count(),
        "results": results
    }
 
    return Response(response_data)
 


